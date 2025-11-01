from flask import Blueprint, jsonify, request, send_file
from models import db, get_dashboard_stats, get_district_progress, get_indicator_performance, get_recent_activities, get_facilities_overview
from models import Participant, Assessment
from utils import ExcelGenerator
from datetime import datetime
import tempfile
import os

# Create blueprint for admin routes
admin_bp = Blueprint('admin', __name__, url_prefix='/api/admin')


@admin_bp.route('/stats', methods=['GET'])
def get_stats():
    """Get dashboard statistics"""
    try:
        stats = get_dashboard_stats()
        return jsonify(stats), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/district-progress', methods=['GET'])
def get_progress():
    """Get district progress data for charts"""
    try:
        day = request.args.get('day', 'total')
        data = get_district_progress(day)
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/indicator-performance', methods=['GET'])
def get_performance():
    """Get indicator performance data"""
    try:
        data = get_indicator_performance()
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/activities', methods=['GET'])
def get_activities():
    """Get recent activities for feed"""
    try:
        limit = request.args.get('limit', 10, type=int)
        activities = get_recent_activities(limit)
        return jsonify(activities), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/facilities', methods=['GET'])
def get_facilities():
    """Get facilities overview"""
    try:
        district = request.args.get('district')
        status = request.args.get('status')
        search = request.args.get('search')
        
        facilities = get_facilities_overview(district, status, search)
        return jsonify(facilities), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/export/participants', methods=['GET'])
def export_participants():
    """Export all participants to Excel"""
    try:
        from flask import session
        from models import log_activity
        
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        query = Participant.query
        
        if start_date:
            query = query.filter(Participant.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query = query.filter(Participant.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
        
        participants = query.all()
        participants_data = [p.to_dict() for p in participants]
        
        if not participants_data:
            return jsonify({'error': 'No participants found'}), 404
        
        # Log export activity
        log_activity(
            'export',
            session.get('username', 'Unknown'),
            {'reportType': 'All Participants', 'count': len(participants_data), 'dateRange': f'{start_date or "all"} to {end_date or "all"}'},
            'admin',
            None,
            request.remote_addr
        )
        
        # Generate Excel without protection (editable for all users)
        filepath, filename = ExcelGenerator.create_participant_excel(participants_data, protect_sheet=False)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/export/assessments', methods=['GET'])
def export_assessments():
    """Export all assessments to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        query = Assessment.query
        
        if start_date:
            query = query.filter(Assessment.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query = query.filter(Assessment.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
        
        assessments = query.all()
        assessments_data = [a.to_dict() for a in assessments]
        
        if not assessments_data:
            return jsonify({'error': 'No assessments found'}), 404
        
        # Log export activity
        from flask import session
        from models import log_activity
        log_activity(
            'export',
            session.get('username', 'Unknown'),
            {'reportType': 'All Assessments', 'count': len(assessments_data), 'dateRange': f'{start_date or "all"} to {end_date or "all"}'},
            'admin',
            None,
            request.remote_addr
        )
        
        # For now, create a simple Excel file
        # In production, you would use create_assessment_excel for each
        # or create a master aggregated report
        
        import openpyxl
        from openpyxl.styles import Font
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Assessments"
        
        # Headers
        headers = ['Facility', 'District', 'Level', 'Assessor', 'Date', 'Overall Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for idx, assessment in enumerate(assessments_data, 2):
            ws.cell(row=idx, column=1, value=assessment.get('facilityName'))
            ws.cell(row=idx, column=2, value=assessment.get('district'))
            ws.cell(row=idx, column=3, value=assessment.get('facilityLevel'))
            ws.cell(row=idx, column=4, value=assessment.get('assessorName'))
            ws.cell(row=idx, column=5, value=assessment.get('assessmentDate'))
            ws.cell(row=idx, column=6, value=assessment.get('overallScore'))
        
        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'all_assessments_{timestamp}.xlsx'
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/export/analytics', methods=['GET'])
def export_analytics():
    """Export combined analytics report"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        # Query participants
        query_p = Participant.query
        if start_date:
            query_p = query_p.filter(Participant.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query_p = query_p.filter(Participant.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
        participants = query_p.all()
        
        # Query assessments
        query_a = Assessment.query
        if start_date:
            query_a = query_a.filter(Assessment.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query_a = query_a.filter(Assessment.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
        assessments = query_a.all()
        
        # Log export activity
        from flask import session
        from models import log_activity
        log_activity(
            'export',
            session.get('username', 'Unknown'),
            {
                'reportType': 'Combined Analytics Report',
                'participantCount': len(participants),
                'assessmentCount': len(assessments),
                'dateRange': f'{start_date or "all"} to {end_date or "all"}'
            },
            'admin',
            None,
            request.remote_addr
        )
        
        # Create combined Excel workbook
        wb = Workbook()
        
        # Sheet 1: Participants
        ws_p = wb.active
        ws_p.title = "Participants"
        p_headers = ['Name', 'Cadre', 'Facility', 'District', 'Mobile', 'Registration Date', 'Campaign Day']
        for col, header in enumerate(p_headers, 1):
            cell = ws_p.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        for idx, p in enumerate(participants, 2):
            ws_p.cell(row=idx, column=1, value=p.participant_name)
            ws_p.cell(row=idx, column=2, value=p.cadre)
            ws_p.cell(row=idx, column=3, value=p.duty_station)
            ws_p.cell(row=idx, column=4, value=p.district)
            ws_p.cell(row=idx, column=5, value=p.mobile_number)
            ws_p.cell(row=idx, column=6, value=p.registration_date.strftime('%Y-%m-%d') if p.registration_date else '')
            ws_p.cell(row=idx, column=7, value=f"Day {p.campaign_day}" if p.campaign_day else 'N/A')
        
        # Sheet 2: Assessments
        ws_a = wb.create_sheet("Assessments")
        a_headers = ['Facility', 'District', 'Level', 'Assessor', 'Date', 'Overall Score', 'Campaign Day']
        for col, header in enumerate(a_headers, 1):
            cell = ws_a.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        for idx, a in enumerate(assessments, 2):
            ws_a.cell(row=idx, column=1, value=a.facility_name)
            ws_a.cell(row=idx, column=2, value=a.district)
            ws_a.cell(row=idx, column=3, value=a.facility_level)
            ws_a.cell(row=idx, column=4, value=a.assessor_name)
            ws_a.cell(row=idx, column=5, value=a.assessment_date.strftime('%Y-%m-%d') if a.assessment_date else '')
            ws_a.cell(row=idx, column=6, value=f"{a.overall_score:.1f}%" if a.overall_score else 'N/A')
            ws_a.cell(row=idx, column=7, value=f"Day {a.campaign_day}" if a.campaign_day else 'N/A')
        
        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'analytics_report_{timestamp}.xlsx'
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/export/all', methods=['GET'])
def export_all():
    """Export everything (participants + assessments) - redirects to analytics export"""
    # Redirect to the analytics export which does the same thing
    return export_analytics()


# ============= DELETE & RESET OPERATIONS (ADMIN ONLY) =============

@admin_bp.route('/delete/participant/<int:participant_id>', methods=['DELETE'])
def delete_participant(participant_id):
    """Delete a specific participant"""
    try:
        participant = Participant.query.get(participant_id)
        if not participant:
            return jsonify({'success': False, 'message': 'Participant not found'}), 404
        
        name = participant.participant_name
        db.session.delete(participant)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted participant: {name}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/delete/assessment/<int:assessment_id>', methods=['DELETE'])
def delete_assessment(assessment_id):
    """Delete a specific assessment"""
    try:
        assessment = Assessment.query.get(assessment_id)
        if not assessment:
            return jsonify({'success': False, 'message': 'Assessment not found'}), 404
        
        facility = assessment.facility_name
        db.session.delete(assessment)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted assessment for: {facility}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/delete/facility/<string:facility_name>', methods=['DELETE'])
def delete_facility_data(facility_name):
    """Delete all data for a specific facility"""
    try:
        # Delete all participants from this facility
        participants = Participant.query.filter_by(duty_station=facility_name).all()
        for p in participants:
            db.session.delete(p)
        
        # Delete all assessments from this facility
        assessments = Assessment.query.filter_by(facility_name=facility_name).all()
        for a in assessments:
            db.session.delete(a)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted all data for facility: {facility_name}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/reset/all', methods=['POST'])
def reset_all_data():
    """Reset/clear all data from the database (ADMIN ONLY - DANGEROUS)"""
    try:
        # Count before deletion
        participant_count = Participant.query.count()
        assessment_count = Assessment.query.count()
        activity_count = ActivityLog.query.count()
        
        # Delete all data
        Participant.query.delete()
        Assessment.query.delete()
        ActivityLog.query.delete()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'🗑️ Database reset complete! Deleted: {participant_count} participants, {assessment_count} assessments, {activity_count} activity logs'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/participants/all', methods=['GET'])
def get_all_participants():
    """Get all participants with full details for management"""
    try:
        participants = Participant.query.order_by(Participant.created_at.desc()).all()
        
        result = []
        for p in participants:
            result.append({
                'id': p.id,
                'name': p.participant_name,
                'cadre': p.cadre,
                'facility': p.duty_station,
                'district': p.district,
                'mobile': p.mobile_number,
                'registrationDate': p.registration_date.strftime('%Y-%m-%d') if p.registration_date else 'N/A',
                'campaignDay': p.campaign_day,
                'submittedBy': p.submitted_by,
                'createdAt': p.created_at.isoformat() if p.created_at else 'N/A'
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/assessments/all', methods=['GET'])
def get_all_assessments():
    """Get all assessments with full details for management"""
    try:
        assessments = Assessment.query.order_by(Assessment.created_at.desc()).all()
        
        result = []
        for a in assessments:
            result.append({
                'id': a.id,
                'facilityName': a.facility_name,
                'district': a.district,
                'facilityLevel': a.facility_level,
                'ownership': a.ownership,
                'assessorName': a.assessor_name,
                'assessmentDate': a.assessment_date.strftime('%Y-%m-%d') if a.assessment_date else 'N/A',
                'overallScore': round(a.overall_score, 1) if a.overall_score else 0,
                'campaignDay': a.campaign_day,
                'submittedBy': a.submitted_by,
                'createdAt': a.created_at.isoformat() if a.created_at else 'N/A'
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

