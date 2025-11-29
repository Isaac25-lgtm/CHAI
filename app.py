from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import io
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///campaign_data.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Admin credentials (change these!)
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'admin123'

# Database Model
class Registration(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    participant_name = db.Column(db.String(200), nullable=False)
    cadre = db.Column(db.String(100), nullable=False)
    district = db.Column(db.String(100), nullable=False)
    facility = db.Column(db.String(200), nullable=False)
    registration_date = db.Column(db.Date, nullable=False)
    day1 = db.Column(db.Boolean, default=False)
    day2 = db.Column(db.Boolean, default=False)
    mobile_number = db.Column(db.String(15), nullable=False)
    mm_registered_names = db.Column(db.String(200), nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'participant_name': self.participant_name,
            'cadre': self.cadre,
            'district': self.district,
            'facility': self.facility,
            'registration_date': self.registration_date.strftime('%Y-%m-%d'),
            'day1': self.day1,
            'day2': self.day2,
            'mobile_number': self.mobile_number,
            'mm_registered_names': self.mm_registered_names,
            'submitted_at': self.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
        }

class Settings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.String(200), nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Create tables
with app.app_context():
    db.create_all()
    # Initialize default settings if they don't exist
    if not Settings.query.filter_by(key='campaign_days').first():
        default_settings = Settings(key='campaign_days', value='1')
        db.session.add(default_settings)
    if not Settings.query.filter_by(key='activity_name').first():
        default_settings = Settings(key='activity_name', value='EDIL Assessment')
        db.session.add(default_settings)
    if not Settings.query.filter_by(key='activity_dates').first():
        default_settings = Settings(key='activity_dates', value='30th November to 7th December 2025')
        db.session.add(default_settings)
    db.session.commit()

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    # Get all settings
    campaign_days_setting = Settings.query.filter_by(key='campaign_days').first()
    activity_name_setting = Settings.query.filter_by(key='activity_name').first()
    activity_dates_setting = Settings.query.filter_by(key='activity_dates').first()
    
    campaign_days = int(campaign_days_setting.value) if campaign_days_setting else 1
    activity_name = activity_name_setting.value if activity_name_setting else 'EDIL Assessment'
    activity_dates = activity_dates_setting.value if activity_dates_setting else '30th November to 7th December 2025'
    
    return render_template('registration.html', 
                         campaign_days=campaign_days,
                         activity_name=activity_name,
                         activity_dates=activity_dates)

@app.route('/api/settings/campaign-days')
def get_campaign_days():
    """API endpoint to get campaign days setting"""
    settings = Settings.query.filter_by(key='campaign_days').first()
    campaign_days = int(settings.value) if settings else 1
    return jsonify({'campaign_days': campaign_days})

@app.route('/submit', methods=['POST'])
def submit_registration():
    try:
        data = request.get_json()
        
        # Parse registration date
        reg_date = datetime.strptime(data.get('registration_date'), '%Y-%m-%d').date()
        
        registration = Registration(
            participant_name=data.get('participant_name'),
            cadre=data.get('cadre'),
            district=data.get('district'),
            facility=data.get('facility'),
            registration_date=reg_date,
            day1=data.get('day1', False),
            day2=data.get('day2', False),
            mobile_number=data.get('mobile_number'),
            mm_registered_names=data.get('mm_registered_names')
        )
        db.session.add(registration)
        db.session.commit()
        
        # Return registration data as JSON for client-side storage
        return jsonify({
            'success': True,
            'data': registration.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/submit/bulk', methods=['POST'])
def submit_bulk_registration():
    try:
        data = request.get_json()
        participants = data.get('participants', [])
        
        if not participants:
            return jsonify({
                'success': False,
                'error': 'No participants provided'
            }), 400
        
        facility = participants[0].get('facility') if participants else 'Unknown'
        registrations = []
        
        for participant_data in participants:
            # Parse registration date
            reg_date = datetime.strptime(participant_data.get('registration_date'), '%Y-%m-%d').date()
            
            registration = Registration(
                participant_name=participant_data.get('participant_name'),
                cadre=participant_data.get('cadre'),
                district=participant_data.get('district'),
                facility=participant_data.get('facility'),
                registration_date=reg_date,
                day1=participant_data.get('day1', False),
                day2=participant_data.get('day2', False),
                mobile_number=participant_data.get('mobile_number'),
                mm_registered_names=participant_data.get('mm_registered_names')
            )
            db.session.add(registration)
            registrations.append(registration)
        
        db.session.commit()
        
        # Return all registration data
        return jsonify({
            'success': True,
            'facility': facility,
            'count': len(registrations),
            'data': [reg.to_dict() for reg in registrations]
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/download/facility/<facility_name>')
def download_facility_data(facility_name):
    try:
        # Get all registrations for this facility
        registrations = Registration.query.filter(
            Registration.facility == facility_name
        ).order_by(Registration.submitted_at.desc()).all()
        
        if not registrations:
            return jsonify({
                'success': False,
                'error': 'No data found for this facility'
            }), 404
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facility Data"
        
        # Define styles
        header_fill = PatternFill(start_color="2B5097", end_color="2B5097", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_font = Font(name="Calibri", size=11)
        cell_alignment = Alignment(vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Headers
        headers = ['No.', "Participant's Name", 'Cadre', 'Duty Station (Facility)', 'District', 
                   'Mobile Number Registered', 'Names Registered on Mobile Money (First & Last Names)',
                   'Day 1', 'Day 2', 'Registration Date']
        
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set row height for header
        ws.row_dimensions[1].height = 35
        
        # Add data
        for idx, reg in enumerate(registrations, start=1):
            row = [
                idx,
                reg.participant_name,
                reg.cadre,
                reg.facility,
                reg.district,
                reg.mobile_number,
                reg.mm_registered_names,
                '✓' if reg.day1 else '✗',
                '✓' if reg.day2 else '✗',
                reg.registration_date.strftime('%Y-%m-%d')
            ]
            ws.append(row)
            
            # Style data cells
            row_num = idx + 1
            for col_num, cell in enumerate(ws[row_num], start=1):
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = cell_alignment
                
                # Center align number column
                if col_num == 1:
                    cell.alignment = center_alignment
                
                # Special styling for Day 1 and Day 2 columns
                if col_num in [8, 9]:
                    cell.alignment = center_alignment
                    cell.font = Font(name="Calibri", size=14, bold=True, 
                                   color="008000" if cell.value == '✓' else "FF0000")
        
        # Set column widths
        column_widths = [6, 28, 20, 35, 20, 25, 45, 10, 10, 18]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create filename
        safe_facility_name = facility_name.replace(' ', '_').replace('/', '_')
        filename = f'{safe_facility_name}_registrations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid credentials. Please try again.', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    # Get filter parameters
    search = request.args.get('search', '')
    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = Registration.query
    
    # Apply filters
    if search:
        query = query.filter(
            db.or_(
                Registration.participant_name.ilike(f'%{search}%'),
                Registration.mobile_number.ilike(f'%{search}%'),
                Registration.facility.ilike(f'%{search}%')
            )
        )
    
    if district:
        query = query.filter(Registration.district == district)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Registration.registration_date >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Registration.registration_date <= date_to_obj)
        except:
            pass
    
    registrations = query.order_by(Registration.submitted_at.desc()).all()
    
    # Get all unique districts for filter dropdown
    districts = db.session.query(Registration.district).distinct().order_by(Registration.district).all()
    districts = [d[0] for d in districts]
    
    # Stats
    total_count = Registration.query.count()
    today = datetime.utcnow().date()
    today_count = Registration.query.filter(
        db.func.date(Registration.submitted_at) == today
    ).count()
    
    # Filtered count
    filtered_count = len(registrations)
    
    return render_template('admin_dashboard.html', 
                         registrations=registrations,
                         total_count=total_count,
                         today_count=today_count,
                         filtered_count=filtered_count,
                         search=search,
                         districts=districts,
                         selected_district=district,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/admin/download/excel')
@login_required
def download_excel():
    # Get same filters as dashboard
    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query with filters
    query = Registration.query
    
    if district:
        query = query.filter(Registration.district == district)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Registration.registration_date >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Registration.registration_date <= date_to_obj)
        except:
            pass
    
    registrations = query.order_by(Registration.submitted_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    
    # Define styles
    header_fill = PatternFill(start_color="2B5097", end_color="2B5097", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Headers
    headers = ['No.', "Participant's Name", 'Cadre', 'Duty Station (Facility)', 'District', 
               'Mobile Number Registered', 'Names Registered on Mobile Money (First & Last Names)',
               'Day 1', 'Day 2', 'Registration Date']
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set row height for header
    ws.row_dimensions[1].height = 35
    
    # Add data
    for idx, reg in enumerate(registrations, start=1):
        row = [
            idx,
            reg.participant_name,
            reg.cadre,
            reg.facility,
            reg.district,
            reg.mobile_number,
            reg.mm_registered_names,
            '✓' if reg.day1 else '✗',
            '✓' if reg.day2 else '✗',
            reg.registration_date.strftime('%Y-%m-%d')
        ]
        ws.append(row)
        
        # Style data cells
        row_num = idx + 1
        for col_num, cell in enumerate(ws[row_num], start=1):
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment
            
            # Center align number column
            if col_num == 1:
                cell.alignment = center_alignment
            
            # Special styling for Day 1 and Day 2 columns
            if col_num in [8, 9]:
                cell.alignment = center_alignment
                cell.font = Font(name="Calibri", size=14, bold=True, 
                               color="008000" if cell.value == '✓' else "FF0000")
    
    # Set column widths
    column_widths = [6, 28, 20, 35, 20, 25, 45, 10, 10, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create filename with filter info
    filename_parts = ['participant_registrations_RAW']
    if district:
        filename_parts.append(district.replace(' ', '_'))
    if date_from or date_to:
        if date_from and date_to:
            filename_parts.append(f'{date_from}_to_{date_to}')
        elif date_from:
            filename_parts.append(f'from_{date_from}')
        elif date_to:
            filename_parts.append(f'until_{date_to}')
    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = '_'.join(filename_parts) + '.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/download/analyzed')
@login_required
def download_analyzed():
    # Get same filters as dashboard
    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query with filters
    query = Registration.query
    
    if district:
        query = query.filter(Registration.district == district)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Registration.registration_date >= date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Registration.registration_date <= date_to_obj)
        except:
            pass
    
    registrations = query.order_by(Registration.submitted_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analyzed Attendance"
    
    # Define styles
    header_fill = PatternFill(start_color="2B5097", end_color="2B5097", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Headers for ANALYZED data (no Day 1/Day 2 columns, just Attendance Date)
    headers = ['No.', "Participant's Name", 'Cadre', 'Duty Station (Facility)', 'District', 
               'Attendance Date', 'Mobile Number Registered', 
               'Names Registered on Mobile Money (First & Last Names)']
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set row height for header
    ws.row_dimensions[1].height = 35
    
    # Add ANALYZED data - participants appear once or twice based on attendance
    row_num = 1
    for reg in registrations:
        # Registration date is considered Day 2
        registration_date = reg.registration_date
        day1_date = registration_date - timedelta(days=1)
        
        # If Day 1 is checked, add a row with Day 1 date
        if reg.day1:
            row_num += 1
            row = [
                row_num - 1,
                reg.participant_name,
                reg.cadre,
                reg.facility,
                reg.district,
                day1_date.strftime('%Y-%m-%d'),
                reg.mobile_number,
                reg.mm_registered_names
            ]
            ws.append(row)
            
            # Style data cells
            for col_num, cell in enumerate(ws[row_num], start=1):
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = cell_alignment
                
                # Center align number and date columns
                if col_num in [1, 6]:
                    cell.alignment = center_alignment
        
        # If Day 2 is checked, add a row with Day 2 date (registration date)
        if reg.day2:
            row_num += 1
            row = [
                row_num - 1,
                reg.participant_name,
                reg.cadre,
                reg.facility,
                reg.district,
                registration_date.strftime('%Y-%m-%d'),
                reg.mobile_number,
                reg.mm_registered_names
            ]
            ws.append(row)
            
            # Style data cells
            for col_num, cell in enumerate(ws[row_num], start=1):
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = cell_alignment
                
                # Center align number and date columns
                if col_num in [1, 6]:
                    cell.alignment = center_alignment
    
    # Set column widths
    column_widths = [6, 28, 20, 35, 20, 18, 25, 45]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create filename with filter info
    filename_parts = ['participant_attendance_ANALYZED']
    if district:
        filename_parts.append(district.replace(' ', '_'))
    if date_from or date_to:
        if date_from and date_to:
            filename_parts.append(f'{date_from}_to_{date_to}')
        elif date_from:
            filename_parts.append(f'from_{date_from}')
        elif date_to:
            filename_parts.append(f'until_{date_to}')
    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = '_'.join(filename_parts) + '.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/delete/<int:id>', methods=['POST'])
@login_required
def delete_registration(id):
    # Preserve filters
    search = request.args.get('search', '')
    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    registration = Registration.query.get_or_404(id)
    db.session.delete(registration)
    db.session.commit()
    flash('Registration deleted successfully.', 'success')
    
    # Redirect with preserved filters
    return redirect(url_for('admin_dashboard', 
                           search=search, 
                           district=district, 
                           date_from=date_from, 
                           date_to=date_to))

@app.route('/admin/clear-all', methods=['POST'])
@login_required
def clear_all():
    # Preserve filters
    search = request.args.get('search', '')
    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    Registration.query.delete()
    db.session.commit()
    flash('All registrations cleared.', 'success')
    
    # Redirect with preserved filters
    return redirect(url_for('admin_dashboard', 
                           search=search, 
                           district=district, 
                           date_from=date_from, 
                           date_to=date_to))

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
def admin_settings():
    if request.method == 'POST':
        # Update campaign days
        campaign_days = request.form.get('campaign_days', '1')
        try:
            campaign_days_int = int(campaign_days)
            if campaign_days_int < 1 or campaign_days_int > 7:
                flash('Campaign days must be between 1 and 7', 'error')
                return redirect(url_for('admin_settings'))
        except ValueError:
            flash('Invalid number of campaign days', 'error')
            return redirect(url_for('admin_settings'))
        
        # Update activity name
        activity_name = request.form.get('activity_name', 'EDIL Assessment').strip()
        if not activity_name:
            activity_name = 'EDIL Assessment'
        
        # Update activity dates
        activity_dates = request.form.get('activity_dates', '30th November to 7th December 2025').strip()
        if not activity_dates:
            activity_dates = '30th November to 7th December 2025'
        
        # Save campaign days
        settings = Settings.query.filter_by(key='campaign_days').first()
        if settings:
            settings.value = str(campaign_days_int)
            settings.updated_at = datetime.utcnow()
        else:
            settings = Settings(key='campaign_days', value=str(campaign_days_int))
            db.session.add(settings)
        
        # Save activity name
        settings = Settings.query.filter_by(key='activity_name').first()
        if settings:
            settings.value = activity_name
            settings.updated_at = datetime.utcnow()
        else:
            settings = Settings(key='activity_name', value=activity_name)
            db.session.add(settings)
        
        # Save activity dates
        settings = Settings.query.filter_by(key='activity_dates').first()
        if settings:
            settings.value = activity_dates
            settings.updated_at = datetime.utcnow()
        else:
            settings = Settings(key='activity_dates', value=activity_dates)
            db.session.add(settings)
        
        db.session.commit()
        flash('Settings updated successfully!', 'success')
        return redirect(url_for('admin_settings'))
    
    # GET request - show settings page
    campaign_days_setting = Settings.query.filter_by(key='campaign_days').first()
    activity_name_setting = Settings.query.filter_by(key='activity_name').first()
    activity_dates_setting = Settings.query.filter_by(key='activity_dates').first()
    
    campaign_days = int(campaign_days_setting.value) if campaign_days_setting else 1
    activity_name = activity_name_setting.value if activity_name_setting else 'EDIL Assessment'
    activity_dates = activity_dates_setting.value if activity_dates_setting else '30th November to 7th December 2025'
    
    return render_template('admin_settings.html', 
                         campaign_days=campaign_days,
                         activity_name=activity_name,
                         activity_dates=activity_dates)

@app.route('/healthz')
def healthz():
    return jsonify(status='ok')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
